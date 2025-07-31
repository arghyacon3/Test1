from pypdf import PdfReader as PyPDFReader
from PyPDF2 import PdfReader as PyPDF2Reader
import pandas as pd
from openpyxl import Workbook

class PDFCommentExtractor:
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.comments_data = []

    def extract_with_pypdf(self):
        """Extract comments using pypdf (for CSV saving)"""
        reader = PyPDFReader(self.pdf_path)
        data = []

        for page_num, page in enumerate(reader.pages, start=1):
            if "/Annots" in page:
                annotations = page["/Annots"]
                for annot_ref in annotations:
                    annot = annot_ref.get_object()
                    content = annot.get("/Contents")
                    if content:
                        author = annot.get("/T", "")
                        data.append({
                            "PageNo": page_num,
                            "Author": author,
                            "Comment": content
                        })
        self.comments_data = data

    def save_to_csv(self, output_path: str):
        if not self.comments_data:
            raise ValueError("No comments to save. Did you run extract_with_pypdf()?")
        df = pd.DataFrame(self.comments_data, columns=["PageNo", "Author", "Comment"])
        df.to_csv(output_path, index=False, encoding='utf-8')
        print(f"✅ CSV saved at: {output_path}")

    def extract_with_pypdf2(self):
        """Extract comments using PyPDF2 (for Excel saving)"""
        reader = PyPDF2Reader(self.pdf_path)
        data = []
        comment_count = 1

        for page_number, page in enumerate(reader.pages, start=1):
            if "/Annots" in page:
                annotations = page["/Annots"]
                for annot in annotations:
                    obj = annot.get_object()
                    if "/Contents" in obj:
                        comment_text = obj["/Contents"]
                        subtext = obj.get("/Subj") or obj.get("/T") or obj.get("/NM") or ""
                        print(
                            f"{'✅' if subtext else '❌'} Subtext for comment #{comment_count} on page {page_number}: {subtext}"
                        )
                        data.append((comment_count, comment_text, page_number, subtext))
                        comment_count += 1

        self.comments_data = data

    def save_to_excel(self, excel_path: str, start_row: int = 10):
        if not self.comments_data:
            raise ValueError("No comments to save. Did you run extract_with_pypdf2()?")

        wb = Workbook()
        ws = wb.active

        # Write headers
        ws["C9"] = "Comment Number"
        ws["D9"] = "Comment Text"
        ws["E9"] = "Page Number"
        ws["F9"] = "Subtext"

        for i, (c_no, text, page_num, subtext) in enumerate(self.comments_data):
            row = start_row + i
            ws[f"C{row}"] = c_no
            ws[f"D{row}"] = text
            ws[f"E{row}"] = page_num
            ws[f"F{row}"] = subtext

        wb.save(excel_path)
        print(f"✅ Excel saved at: {excel_path}")
